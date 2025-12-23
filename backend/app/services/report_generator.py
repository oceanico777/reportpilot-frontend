from sqlalchemy.orm import Session
from .. import models
import time

import pytesseract
from PIL import Image
from pypdf import PdfReader
import magic
import os
import io
import platform
from datetime import datetime

# Set Tesseract path for Windows
if platform.system() == "Windows":
    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

def extract_text_from_pdf(content: bytes) -> str:
    try:
        reader = PdfReader(io.BytesIO(content))
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        print(f"Error extracting text from PDF: {e}")
        return ""

def process_text_with_llm(text: str):
    # Placeholder for LLM processing
    # In a real app, this would send 'text' to OpenAI/Gemini to extract structured data
    return f"Processed Text (Length: {len(text)} chars). Content Preview: {text[:100]}..."

def create_report(report_id: str):
    """
    Generate report details using Gemini OCR (background task)
    """
    from ..database import SessionLocal
    from ..services.storage import storage_service
    from ..services.ocr import process_receipt_with_gemini
    
    db = SessionLocal()
    report = db.query(models.Report).filter(models.Report.id == report_id).first()
    if not report:
        db.close()
        return

    try:
        # 1. Get file content from Supabase
        file_path = report.source_file_path
        if not file_path:
             raise Exception("No source file path provided")
             
        contents = storage_service.download_file(file_path)
        if not contents:
            # Fallback to local for dev/legacy
            if os.path.exists(file_path):
                with open(file_path, "rb") as f:
                    contents = f.read()
            else:
                raise Exception(f"File not found in Supabase or locally: {file_path}")

        # 2. Process with Gemini
        extracted_data = process_receipt_with_gemini(contents)
        
        # 3. Update Report with extracted data
        vendor = extracted_data.get('vendor', "Comercio no detectado")
        amount = float(extracted_data.get('amount', 0.0))
        currency = extracted_data.get('currency', 'COP')
        category = extracted_data.get('category', '📦 Otros')
        
        summary = (
            f"Factura de {vendor}. "
            f"Total: {currency} {amount}. "
            f"Categoría: {category}"
        )
        
        report.vendor = vendor
        report.vendor_nit = extracted_data.get('vendor_nit')
        report.amount = amount
        report.currency = currency
        report.category = category
        report.summary_text = summary
        report.status = models.ReportStatus.PENDING_REVIEW.value
        
    except Exception as e:
        import logging
        logging.error(f"Error generating report {report_id}: {e}")
        report.status = models.ReportStatus.FAILED.value
        report.summary_text = f"Error: {str(e)}"
    
    finally:
        db.commit()
        db.close()

def generate_tour_summary(reports):
    """
    Aggregates reports by Tour ID and Category for the Admin View.
    Returns a dictionary structure.
    """
    tour_map = {}
    
    for r in reports:
        tid = r.tour_id or "Unassigned"
        if tid not in tour_map:
            tour_map[tid] = {
                "tour_id": tid,
                "total_amount": 0.0,
                "reports_count": 0,
                "categories": {},
                "guide_id": r.user_id,
                "guide_name": r.client_name,
                "has_duplicates": False
            }
        
        if r.is_duplicate:
            tour_map[tid]["has_duplicates"] = True

        amount = r.amount or 0
        cat = r.category or "📦 Otros"
        
        tour_map[tid]["total_amount"] += amount
        tour_map[tid]["reports_count"] += 1
        tour_map[tid]["categories"][cat] = tour_map[tid]["categories"].get(cat, 0) + amount

    return list(tour_map.values())

def generate_excel_report(reports):
    """
    Generates a professional Excel report with Financial Summary and Details.
    Returns bytes buffer.
    """
    import pandas as pd
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # 1. Prepare Data & Calculations
    data = []
    
    total_advances = 0
    total_collections = 0
    total_expenses = 0
    
    # Financial Categories mapping
    INCOME_CATEGORIES = ["ANTICIPO_RECIBIDO", "RECAUDO_CLIENTE"]
    
    category_totals = {} # For detailed category breakdown
    total_general = 0 # Sum of all amounts
    
    for r in reports:
        raw_cat = r.category or "OTROS"
        cat = raw_cat.upper().replace("Ó", "O").replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ú", "U")
        
        # Normalize simple mapping
        cat_key = cat # Default
        if "ALIMENTACION" in cat or "RESTAURANTE" in cat or "ALMUERZO" in cat: cat_key = "ALIMENTACION"
        elif "TRANSPORTE" in cat or "TAXI" in cat or "UBER" in cat: cat_key = "TRANSPORTE"
        elif "ENTRADA" in cat: cat_key = "ENTRADAS"
        elif "PARQUEADERO" in cat: cat_key = "PARQUEADERO"
        elif "REFRIGERIO" in cat: cat_key = "REFRIGERIO"
        elif "PROPINA" in cat: cat_key = "PROPINAS"
        
        category_totals[cat_key] = category_totals.get(cat_key, 0) + (r.amount or 0)
        total_general += (r.amount or 0)           
        
        amount = r.amount or 0
        
        # Financial Sums (using original category for these specific checks)
        if raw_cat == "ANTICIPO_RECIBIDO":
            total_advances += amount
        elif raw_cat == "RECAUDO_CLIENTE":
            total_collections += amount
        else:
            total_expenses += amount
            
        data.append({
            "ID Reporte": r.id,
            "Fecha": r.created_at.strftime("%d/%m/%Y") if r.created_at else "",
            "Tour ID": r.tour_id or "Sin Asignar",
            "Guía": r.client_name or "Desconocido",
            "Proveedor": r.vendor,
            "NIT": r.vendor_nit or "No Detectado", # NEW
            "Categoría": cat,
            "Descripción": r.summary_text,
            "Monto (COP)": amount,
            "Estado": r.status or "BORRADOR"
        })
        
    df = pd.DataFrame(data)

    # Calculate Net Balance (Liability)
    # Logic: Liability = (Advances + Collections) - Expenses
    # If Positive: Guide owes agency (Reintegrar)
    # If Negative: Agency owes guide (Reembolsar)
    net_balance = (total_advances + total_collections) - total_expenses
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # --- SHEET 1: LIQUIDATION SUMMARY ---
        # Create a summary dataframe structure
        summary_data = {
            "Concepto": [
                "TOTAL ANTICIPOS RECIBIDOS",
                "TOTAL RECAUDOS CLIENTE",
                "SUBTOTAL INGRESOS (Responsabilidad)",
                "TOTAL GASTOS LEGALIZADOS",
                "BALANCE NETO (A LIQUIDAR)"
            ],
            "Monto": [
                total_advances,
                total_collections,
                total_advances + total_collections,
                total_expenses,
                net_balance
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Resumen Liquidación', startrow=3, index=False)
        
        workbook = writer.book
        ws_summary = writer.sheets['Resumen Liquidación']
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Slate
        currency_format = '"$" #,##0'
        border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Title
        ws_summary['A1'] = "LIQUIDACIÓN DE GASTOS DE VIAJE"
        ws_summary['A1'].font = Font(bold=True, size=16, color="0F172A")
        ws_summary.merge_cells('A1:B1')
        
        # Tour Info
        if not df.empty:
            tour_id = df.iloc[0]["Tour ID"]
            guide_name = df.iloc[0]["Guía"]
            ws_summary['A2'] = f"Tour ID: {tour_id} | Guía: {guide_name}"
        else:
            ws_summary['A2'] = "Sin datos de reporte"
            
        ws_summary['A2'].font = Font(bold=True, italic=True, size=11, color="64748B")

        # Format Summary Table
        # Headers (Row 4)
        for cell in ws_summary[4]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data Rows (Row 5 to 9)
        for row in ws_summary.iter_rows(min_row=5, max_row=9, min_col=1, max_col=2):
            for cell in row:
                cell.border = border_style
                if cell.column == 2: # Amount Column
                    cell.number_format = currency_format
                    cell.font = Font(name='Courier New', bold=True)
        
        # Highlight Net Balance (Last Row)
        last_row = 9
        balance_cell = ws_summary.cell(row=last_row, column=2)
        if net_balance > 0:
            status_text = "⚠️ EL GUÍA DEBE REINTEGRAR"
            balance_cell.font = Font(bold=True, color="DC2626") # Red
        elif net_balance < 0:
            status_text = "🔵 LA AGENCIA DEBE REEMBOLSAR"
            balance_cell.font = Font(bold=True, color="2563EB") # Blue
        else:
            status_text = "✅ CUADRE PERFECTO"
            balance_cell.font = Font(bold=True, color="16A34A") # Green
            
        ws_summary.cell(row=last_row, column=3, value=status_text).font = Font(bold=True)

        # Warning Note for Unbalanced Closure
        if abs(net_balance) > 1000:
            ws_summary.cell(row=last_row + 2, column=1, value="⚠️ NOTA: Liquidación con Diferencia Pendiente (> $1.000 COP)").font = Font(bold=True, color="DC2626", size=11)

        # Adjust Column Widths
        ws_summary.column_dimensions['A'].width = 40
        ws_summary.column_dimensions['B'].width = 25
        ws_summary.column_dimensions['C'].width = 35

        # --- SHEET 2: DETAILED REPORTS ---
        if not df.empty:
            df.to_excel(writer, sheet_name='Detalle Movimientos', index=False, startrow=1)
            ws_details = writer.sheets['Detalle Movimientos']
            
            # Title
            ws_details['A1'] = "DETALLE DE MOVIMIENTOS"
            ws_details['A1'].font = Font(bold=True, size=14)
            
            # Headers Customization
            for cell in ws_details[2]: # Row 2 because startrow=1
                cell.font = header_font
                cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid") # Lighter Slate
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-filter
            ws_details.auto_filter.ref = ws_details.dimensions
            
            # Formatting Columns
            for row in ws_details.iter_rows(min_row=3):
                currency_cell = row[8] # 'Monto (COP)' is now 9th column (0-indexed 8) because of NIT
                currency_cell.number_format = currency_format
                
            # Adjust Widths
            for col in ws_details.columns:
                max_length = 0
                column = [cell for cell in col]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) if max_length < 50 else 50
                ws_details.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    output.seek(0)
    return output

def generate_clearance_act(tour_data, signature_path):
    """
    Generates a PDF 'Acta de Liquidación' using WeasyPrint.
    Returns bytes buffer.
    """
    # 1. Prepare Data
    company_name = tour_data.get("company_name", "Empresa de Turismo")
    tour_id = tour_data.get("tour_id", "Unknown")
    guide_name = tour_data.get("guide_name", "Unknown")
    closed_at = tour_data.get("closed_at", datetime.now()).strftime("%d/%m/%Y %H:%M")
    
    advances = tour_data.get("total_advances", 0)
    collections = tour_data.get("total_collections", 0)
    expenses = tour_data.get("total_expenses", 0)
    balance = (advances + collections) - expenses
    
    balance_text = ""
    status_icon = ""
    balance_bg = "#ecfdf5"
    balance_color = "#10b981"
    
    if balance > 0:
        balance_text = f"EL GUÍA DEBE REINTEGRAR: ${balance:,.0f} COP"
        balance_color = "#ef4444" # Red 500
        balance_bg = "#fef2f2"
        status_icon = "⚠️"
    elif balance < 0:
        balance_text = f"LA AGENCIA DEBE REEMBOLSAR: ${abs(balance):,.0f} COP"
        balance_color = "#3b82f6" # Blue 500
        balance_bg = "#eff6ff"
        status_icon = "🔵"
    else:
        balance_text = "PAZ Y SALVO (Saldo $0)"
        balance_color = "#10b981" # Emerald 500
        balance_bg = "#ecfdf5"
        status_icon = "✅"

    # 2. HTML Template
    expense_rows = ""
    for exp in tour_data.get("expense_details", []):
        expense_rows += f"""
        <tr>
            <td style="color: #64748b;">{exp['date']}</td>
            <td style="font-weight: 500; color: #1e293b;">{exp['vendor']}</td>
            <td style="color: #64748b; font-family: monospace;">{exp['vendor_nit']}</td>
            <td><span class="badge">{exp['category']}</span></td>
            <td style="text-align: right; font-weight: 600; font-family: 'Courier New', monospace;">${exp['amount']:,.0f}</td>
        </tr>
        """

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{ margin: 0; size: A4; }}
            body {{ 
                font-family: 'Inter', system-ui, -apple-system, sans-serif; 
                color: #334155; 
                line-height: 1.5; 
                font-size: 11px; 
                margin: 0;
                padding: 40px;
                background: #fff;
            }}
            
            /* Branding Header */
            .header {{ 
                display: flex; 
                justify-content: space-between; 
                align-items: center;
                border-bottom: 2px solid #e2e8f0; 
                padding-bottom: 20px; 
                margin-bottom: 30px; 
            }}
            .logo-text {{
                font-size: 24px;
                font-weight: 800;
                color: #0f172a;
                letter-spacing: -0.5px;
            }}
            .logo-sub {{
                color: #64748b;
                font-size: 10px;
                font-weight: 500;
                text-transform: uppercase;
                letter-spacing: 1px;
            }}
            .document-title {{
                text-align: right;
            }}
            .main-title {{
                font-size: 18px;
                font-weight: 700;
                color: #0f172a;
                text-transform: uppercase;
            }}
            .doc-id {{
                font-family: monospace;
                color: #64748b;
                font-size: 11px;
                margin-top: 4px;
            }}

            /* Info grid */
            .info-grid {{
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 20px;
                margin-bottom: 30px;
                background: #f8fafc;
                padding: 15px;
                border-radius: 8px;
                border: 1px solid #e2e8f0;
            }}
            .info-item label {{
                display: block;
                font-size: 9px;
                text-transform: uppercase;
                letter-spacing: 0.5px;
                color: #64748b;
                margin-bottom: 2px;
                font-weight: 600;
            }}
            .info-item div {{
                font-size: 12px;
                font-weight: 600;
                color: #0f172a;
            }}

            /* Tables */
            table {{ width: 100%; border-collapse: collapse; }}
            
            .summary-section {{ margin-bottom: 30px; }}
            .section-title {{
                font-size: 12px;
                font-weight: 700;
                color: #0f172a;
                text-transform: uppercase;
                letter-spacing: 0.5px;
                border-left: 4px solid #2563eb; /* Brand Blue */
                padding-left: 10px;
                margin-bottom: 15px;
            }}

            .summary-card {{
                background: #fff;
                border-radius: 8px;
                overflow: hidden;
                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
                border: 1px solid #e2e8f0;
            }}
            .summary-table th {{
                background: #1e293b; /* Dark Slate */
                color: #fff;
                font-weight: 600;
                padding: 10px 15px;
                font-size: 10px;
                text-transform: uppercase;
                text-align: left;
            }}
            .summary-table td {{
                padding: 10px 15px;
                border-bottom: 1px solid #f1f5f9;
                font-size: 11px;
            }}
            .summary-table tr:last-child td {{ border-bottom: none; }}
            
            .total-row {{
                background-color: #f8fafc;
                font-weight: 700;
                color: #0f172a;
            }}

            /* Balance Card */
            .balance-card {{
                margin-top: 20px;
                background-color: {balance_bg};
                border: 1px solid {balance_color}40;
                border-radius: 8px;
                padding: 20px;
                text-align: center;
            }}
            .balance-label {{
                color: #64748b;
                font-size: 10px;
                text-transform: uppercase;
                font-weight: 600;
                margin-bottom: 5px;
            }}
            .balance-amount {{
                color: {balance_color};
                font-size: 20px;
                font-weight: 800;
            }}
            .status-icon {{ font-size: 16px; vertical-align: middle; margin-right: 5px; }}

            /* Details Table */
            .details-table th {{
                background: #f1f5f9;
                color: #475569;
                font-weight: 600;
                text-transform: uppercase;
                font-size: 10px;
                padding: 8px;
                border-bottom: 2px solid #e2e8f0;
                text-align: left;
            }}
            .details-table td {{
                padding: 8px;
                border-bottom: 1px solid #e2e8f0;
            }}
            .badge {{
                background: #e2e8f0;
                color: #475569;
                padding: 2px 6px;
                border-radius: 4px;
                font-size: 9px;
                font-weight: 600;
            }}

            /* Footer & Signature */
            .bottom-section {{
                margin-top: 40px;
                page-break-inside: avoid;
            }}
            .declaration {{
                font-style: italic;
                color: #64748b;
                font-size: 10px;
                background: #fafafa;
                padding: 15px;
                border-radius: 6px;
                border: 1px dashed #cbd5e1;
                margin-bottom: 30px;
            }}
            .signatures-grid {{
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 40px;
                margin-top: 20px;
            }}
            .sig-box {{
                text-align: center;
            }}
            .sig-line {{
                border-top: 1px solid #0f172a;
                margin-top: 40px;
                margin-bottom: 5px;
            }}
            .sig-name {{ font-weight: 700; color: #0f172a; font-size: 11px; }}
            .sig-role {{ color: #64748b; font-size: 10px; text-transform: uppercase; }}
            
            .sig-img {{
                height: 50px;
                margin-bottom: -40px; /* Overlap line */
                position: relative;
                z-index: 1;
            }}

            .footer-branding {{
                position: fixed;
                bottom: 20px;
                left: 40px;
                right: 40px;
                text-align: center;
                color: #94a3b8;
                font-size: 9px;
                border-top: 1px solid #f1f5f9;
                padding-top: 10px;
            }}
        </style>
    </head>
    <body>
        <!-- Header -->
        <div class="header">
            <div>
                <div class="logo-text">ReportPilot</div>
                <div class="logo-sub">Expense Management</div>
            </div>
            <div class="document-title">
                <div class="main-title">Acta de Liquidación</div>
                <div class="doc-id">#{tour_id}</div>
            </div>
        </div>

        <!-- Info Grid -->
        <div class="info-grid">
            <div class="info-item">
                <label>Responsable / Guía</label>
                <div>{guide_name}</div>
            </div>
            <div class="info-item">
                <label>Fecha de Cierre</label>
                <div>{closed_at}</div>
            </div>
            <div class="info-item">
                <label>Empresa</label>
                <div>{company_name}</div>
            </div>
            <div class="info-item">
                <label>Estado</label>
                <div>Validado</div>
            </div>
        </div>

        <!-- Financial Summary -->
        <div class="summary-section">
            <div class="section-title">Resumen Financiero</div>
            <div class="summary-card">
                <table class="summary-table">
                    <thead>
                        <tr>
                            <th>Concepto</th>
                            <th style="text-align: right;">Monto</th>
                        </tr>
                    </thead>
                    <tbody>
                        <tr>
                            <td>(+) Asignación Inicial (Anticipos)</td>
                            <td style="text-align: right;">${advances:,.0f}</td>
                        </tr>
                        <tr>
                            <td>(+) Recaudos en Ruta</td>
                            <td style="text-align: right;">${collections:,.0f}</td>
                        </tr>
                        <tr class="total-row">
                            <td>(=) Total Responsabilidad</td>
                            <td style="text-align: right;">${(advances + collections):,.0f}</td>
                        </tr>
                        <tr>
                            <td style="color: #ef4444;">(-) Gastos Legalizados</td>
                            <td style="text-align: right; color: #ef4444;">-${expenses:,.0f}</td>
                        </tr>
                    </tbody>
                </table>
            </div>

            <!-- Balance Result -->
            <div class="balance-card">
                <div class="balance-label">Balance Final de Liquidación</div>
                <div class="balance-amount"><span class="status-icon">{status_icon}</span> {balance_text}</div>
            </div>
        </div>

        <!-- Expense Details -->
        <div class="summary-section">
            <div class="section-title">Detalle de Gastos</div>
            <table class="details-table">
                <thead>
                    <tr>
                        <th style="width: 15%;">Fecha</th>
                        <th style="width: 25%;">Proveedor</th>
                        <th style="width: 15%;">NIT</th>
                        <th style="width: 20%;">Categoría</th>
                        <th style="width: 25%; text-align: right;">Monto</th>
                    </tr>
                </thead>
                <tbody>
                    {expense_rows}
                </tbody>
            </table>
        </div>

        <!-- Signatures -->
        <div class="bottom-section">
            <div class="declaration">
                Declaración de conformidad: Certifico que los gastos aquí detallados corresponden a la ejecución real de mis actividades laborales y que los soportes adjuntos son verídicos. Acepto el balance final resultante de esta auditoría automática.
            </div>

            <div class="signatures-grid">
                <div class="sig-box">
                    <img src="{signature_path}" class="sig-img" alt="Firma"/>
                    <div class="sig-line"></div>
                    <div class="sig-name">{guide_name}</div>
                    <div class="sig-role">Firma Responsable</div>
                </div>
                <div class="sig-box">
                    <div style="height: 50px;"></div>
                    <div class="sig-line"></div>
                    <div class="sig-name">ReportPilot</div>
                    <div class="sig-role">Auditoría Automática</div>
                </div>
            </div>
        </div>

        <div class="footer-branding">
            Generated by ReportPilot • {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        </div>
    </body>
    </html>
    """

    # 3. Save Debug HTML (For Development)
    try:
        debug_path = os.path.join("uploads", "debug_acta.html")
        os.makedirs("uploads", exist_ok=True)
        with open(debug_path, "w", encoding="utf-8") as f:
            f.write(html_content)
    except Exception as e:
        print(f"Failed to save debug HTML: {e}")

    import logging
    # 4. Generate PDF using WeasyPrint
    try:
        from weasyprint import HTML
        pdf_file = HTML(string=html_content, base_url=".").write_pdf()
        return pdf_file
    except OSError as e:
        logging.error(f"WeasyPrint not available (GTK missing?): {e}")
        # Return a dummy PDF if WeasyPrint fails
        return b"%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj 2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj 3 0 obj<</Type/Page/MediaBox[0 0 595 842]/Parent 2 0 R/Resources<<>>>>endobj xref\n0 4\n0000000000 65535 f\n0000000010 00000 n\n0000000060 00000 n\n0000000117 00000 n\ntrailer<</Size 4/Root 1 0 R>>startxref\n223\n%%EOF"
    except Exception as e:
        logging.error(f"Error generating PDF: {e}")
        return None
