from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime, date
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
try:
    from weasyprint import HTML, CSS
except OSError:
    # WeasyPrint requires GTK libraries. If missing, PDF export will fail but app should load.
    HTML = None
    CSS = None
    print("Warning: WeasyPrint (GTK) not found. PDF export disabled.")
from jinja2 import Template

from ..database import get_db
from .. import models
from ..auth import get_current_user

router = APIRouter()

# --- HTML TEMPLATE FOR PDF ---
PDF_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <style>
        body { font-family: sans-serif; font-size: 10pt; color: #333; }
        h1 { text-align: center; color: #1a1a1a; margin-bottom: 5px; }
        .subtitle { text-align: center; color: #666; font-size: 9pt; margin-bottom: 20px; }
        table { width: 100%; border-collapse: collapse; margin-top: 10px; }
        th { background-color: #f3f4f6; text-align: left; padding: 8px; font-weight: bold; border-bottom: 2px solid #ddd; font-size: 9pt; }
        td { padding: 8px; border-bottom: 1px solid #eee; font-size: 9pt; }
        .amount { text-align: right; font-family: monospace; }
        .total-row td { border-top: 2px solid #333; font-weight: bold; background-color: #fff; }
        .footer { position: fixed; bottom: 0; width: 100%; text-align: center; font-size: 8pt; color: #999; }
    </style>
</head>
<body>
    <h1>Reporte de Gastos</h1>
    <div class="subtitle">
        Generado para: {{ user_name }}<br>
        Periodo: {{ start_date }} - {{ end_date }}
    </div>

    <table>
        <thead>
            <tr>
                <th style="width: 15%">Fecha</th>
                <th style="width: 25%">Proveedor</th>
                <th style="width: 25%">Categoría</th>
                <th style="width: 20%">Cliente/Tour</th>
                <th style="width: 15%; text-align: right">Monto (COP)</th>
            </tr>
        </thead>
        <tbody>
            {% for r in reports %}
            <tr>
                <td>{{ r.date }}</td>
                <td>{{ r.vendor }}</td>
                <td>{{ r.category }}</td>
                <td>{{ r.tour }}</td>
                <td class="amount">{{ r.amount }}</td>
            </tr>
            {% endfor %}
            <tr class="total-row">
                <td colspan="4" style="text-align: right">TOTAL</td>
                <td class="amount">{{ total_amount }}</td>
            </tr>
        </tbody>
    </table>

    <div class="footer">
        Generado por ReportPilot - {{ now }}
    </div>
</body>
</html>
"""

def format_currency(amount):
    return "${:,.0f}".format(amount).replace(",", ".")

from ..auth import get_current_user, get_user_company

@router.get("/gastos")
def export_gastos(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    format: str = Query(..., regex="^(pdf|xlsx)$"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
    company_id: str = Depends(get_user_company)
):
    # 1. Query Data Scoped to Company
    query = db.query(models.Report).filter(models.Report.company_id == company_id)
    
    if start_date:
        query = query.filter(models.Report.created_at >= start_date)
    if end_date:
        # End of day
        query = query.filter(models.Report.created_at <= datetime.combine(end_date, datetime.max.time()))
        
    reports = query.order_by(models.Report.created_at.desc()).all()
    
    # Pre-process data
    data_list = []
    total = 0
    for r in reports:
        amt = r.amount or 0
        total += amt
        data_list.append({
            "date": r.created_at.strftime("%Y-%m-%d"),
            "vendor": r.vendor or "N/A",
            "category": r.category or "Sin Categoría",
            "tour": r.tour_id or "-",
            "amount_raw": amt,
            "amount": format_currency(amt)
        })

    # 2. GENERATE FILE
    
    if format == "pdf":
        if HTML is None:
             raise HTTPException(status_code=400, detail="PDF generation is currently disabled on this server (missing libraries). Please try Excel export.")

        context = {
            "user_name": current_user["full_name"],
            "start_date": start_date or "Inicio",
            "end_date": end_date or "Hoy",
            "reports": data_list,
            "total_amount": format_currency(total),
            "now": datetime.now().strftime("%Y-%m-%d %H:%M")
        }
        
        template = Template(PDF_TEMPLATE)
        html_content = template.render(context)
        
        pdf_file = HTML(string=html_content).write_pdf()
        
        filename = f"Gastos_{start_date or 'Inicio'}_{end_date or 'Hoy'}.pdf"
        return StreamingResponse(
            io.BytesIO(pdf_file),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    elif format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Gastos"
        
        # Headers
        headers = ["Fecha", "Proveedor", "Categoría", "Detalle/Tour", "Monto", "Impuestos", "Total", "Comentarios"]
        ws.append(headers)
        
        # Style Headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Rows
        for item in data_list:
            ws.append([
                item["date"],
                item["vendor"],
                item["category"],
                item["tour"],
                item["amount_raw"], # Monto
                0, # Impuestos (Placeholder)
                item["amount_raw"], # Total
                "" # Comentarios
            ])
            
        # Column Widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Gastos_Export_{start_date or 'All'}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
@router.get("/tours/{tour_id}/xlsx")
def export_tour_xlsx(
    tour_id: str,
    db: Session = Depends(get_db),
    company_id: str = Depends(get_user_company)
):
    # 1. Query Data
    reports = db.query(models.Report).filter(
        models.Report.company_id == company_id,
        models.Report.tour_id == tour_id
    ).filter(
        ~models.Report.category.in_(["ANTICIPO_RECIBIDO", "RECAUDO_CLIENTE"])
    ).all()
    
    # 2. Setup Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Legalización {tour_id}"
    
    # --- STYLES ---
    gold_fill = PatternFill(start_color="C5A658", end_color="C5A658", fill_type="solid") # Gold color from image
    white_font = Font(bold=True, color="FFFFFF")
    black_font_bold = Font(bold=True, color="000000")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    border_thick_bottom = Border(bottom=Side(style='thick'))
    
    # --- HEADER SECTION ---
    # Row 2: Title
    ws.merge_cells('B2:F2')
    cell = ws['B2']
    cell.value = "LEGALIZACIÓN HANSA TOURS"
    cell.font = Font(bold=True, size=14, color="C5A658")
    cell.alignment = Alignment(horizontal="center")
    
    # Row 4-8: Metadata Left
    metadata_labels = [
        ("FECHA", date.today().strftime("%Y-%m-%d")),
        ("NOMBRE DEL CLIENTE Y HOTEL", "Varios / Tour"),
        ("TIPO DE TOUR", tour_id),
        ("SOLICITADO POR", "Operaciones"),
        ("TARIFA DE GUIA", "")
    ]
    
    start_row = 4
    for i, (label, val) in enumerate(metadata_labels):
        row = start_row + i
        # Label
        ws[f'B{row}'] = label
        ws[f'B{row}'].fill = gold_fill
        ws[f'B{row}'].font = black_font_bold
        ws[f'B{row}'].border = border_thin
        # Value
        ws.merge_cells(f'C{row}:D{row}')
        ws[f'C{row}'] = val
        ws[f'C{row}'].border = border_thin
        
    # Row 4-6: Metadata Right (Summary)
    right_labels = ["CANT. HORAS", "CANT. PAX", "ANTICIPO (BASE)", "SALDO DE ANTICIPO"]
    ws[f'E4'] = "CANT. HORAS"
    ws[f'E5'] = "CANT. PAX"
    ws[f'E6'] = "ANTICIPO (BASE)"
    ws[f'E7'] = "SALDO DE ANTICIPO"
    
    for r in range(4, 8):
        ws[f'E{r}'].fill = gold_fill
        ws[f'E{r}'].font = black_font_bold
        ws[f'E{r}'].border = border_thin
        ws[f'F{r}'].border = border_thin
        
    # --- EXPENSES SECTION ---
    # Map categories to hardcoded rows
    # Structure: (Row Label, DB Category Identifier)
    expense_rows_map = [
        ("ENTRADAS", "ENTRADAS"),
        ("PARQUEADERO (bike tours)", "PARQUEADERO"),
        ("REFRIGERIO", "REFRIGERIO"),
        ("ALMUERZO CLIENTES", "ALIMENTACION"),
        ("AUX. ALMUERZO GUÍA", "ALMUERZO_GUIA"),
        ("TAXIS AUTORIZADOS", "TRANSPORTE"),
        ("PROPINAS", "PROPINAS"),
        ("COMEN. TRIPADVISOR (40.000)", "COMISION"),
        ("COMISIÓN VENTA TOURS (5%)", "COMISION_VENTA"),
        ("TOTAL GASTO", "TOTAL")
    ]
    
    # Calculate Totals per Category
    category_totals = {}
    total_general = 0
    
    for r in reports:
        cat = r.category or "OTROS"
        # Normalize simple mapping
        if "ALIMENTACION" in cat: cat_key = "ALIMENTACION"
        elif "TRANSPORTE" in cat: cat_key = "TRANSPORTE"
        else: cat_key = cat
        
        category_totals[cat_key] = category_totals.get(cat_key, 0) + (r.amount or 0)
        total_general += (r.amount or 0)

    # Render Rows
    start_row_expenses = 10
    
    # Header for Expenses
    ws[f'B{start_row_expenses}'] = "CONCEPTO"
    ws[f'C{start_row_expenses}'] = "VALOR"
    ws[f'D{start_row_expenses}'] = "DETALLE / OBSERVACIONES"
    for col in ['B', 'C', 'D']:
        ws[f'{col}{start_row_expenses}'].fill = gold_fill
        ws[f'{col}{start_row_expenses}'].font = black_font_bold
        ws[f'{col}{start_row_expenses}'].alignment = Alignment(horizontal="center")
        
    current_row = start_row_expenses + 1
    
    for label, db_key in expense_rows_map:
        if label == "TOTAL GASTO": continue # Skip, we add at end
        
        # Determine value (simple fuzzy match)
        val = 0
        if db_key == "ALIMENTACION": val = category_totals.get("ALIMENTACION", 0) + category_totals.get("RESTAURANTE", 0)
        elif db_key == "TRANSPORTE": val = category_totals.get("TRANSPORTE", 0) + category_totals.get("TAXI", 0)
        elif db_key == "ENTRADAS": val = category_totals.get("ENTRADAS", 0)
        else: val = category_totals.get(db_key, 0)
        
        ws[f'B{current_row}'] = label
        ws[f'B{current_row}'].fill = gold_fill
        ws[f'B{current_row}'].font = black_font_bold
        ws[f'B{current_row}'].border = border_thin
        
        ws[f'C{current_row}'] = val
        ws[f'C{current_row}'].number_format = '"$"#,##0'
        ws[f'C{current_row}'].border = border_thin
        
        ws[f'D{current_row}'] = "" # Space for notes
        ws[f'D{current_row}'].border = border_thin
        
        current_row += 1
        
    # TOTAL ROW
    ws[f'B{current_row}'] = "TOTAL GASTO"
    ws[f'C{current_row}'] = total_general
    ws[f'C{current_row}'].number_format = '"$"#,##0'
    ws[f'C{current_row}'].font = Font(bold=True)
    
    # --- COLUMNS ---
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Legalizacion_{tour_id}.xlsx"}
    )
